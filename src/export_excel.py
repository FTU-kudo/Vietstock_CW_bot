# ============================================================
# EXPORT_EXCEL.PY
# Tạo 2 file Excel theo đúng layout trong PDF "Bản tin chứng quyền":
#   File 1: KẾT QUẢ GIAO DỊCH TRONG PHIÊN VÀ CÁC CHỈ SỐ LIÊN QUAN
#   File 2: THÔNG TIN CHỨNG QUYỀN
# ============================================================

import pandas as pd
from datetime import date
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── ĐỊNH NGHĨA CỘT THEO ĐÚNG THỨ TỰ TRONG PDF ──────────────

# Bảng 1: "KẾT QUẢ GIAO DỊCH TRONG PHIÊN VÀ CÁC CHỈ SỐ LIÊN QUAN"
TABLE1_COLUMNS = [
    ("ma_cw",                "Mã"),
    ("pct_thay_doi",         "Thay đổi(+/-%)"),
    ("gia_hien_tai",         "Đóng cửa"),
    ("klgd",                 "KLGD (Hợp đồng)"),
    ("gtgd_ty_vnd",          "GTGD (tỷ VNĐ)"),
    ("trang_thai_tien",      "Trạng thái tiền"),
    ("premium",              "Premium"),
    ("don_bay",               "Đòn bẩy"),
    ("so_phien_con_lai",      "Số phiên còn lại"),
    ("do_bien_dong_lich_su",  "Độ biến động ngầm định"),
]

# Bảng 2: "THÔNG TIN CHỨNG QUYỀN"
TABLE2_COLUMNS = [
    ("ma_cw",              "Mã chứng quyền"),
    ("to_chuc_ph_cw",      "TCPH"),
    ("thoi_han",           "Thời hạn"),
    ("ty_le_chuyen_doi",   "Tỷ lệ chuyển đổi"),
    ("gia_phat_hanh",      "Giá phát hành"),
    ("kl_niem_yet",        "Khối lượng phát hành"),
    ("gia_thuc_hien",      "Giá thực hiện"),
    ("ngay_gd_cuoi_cung",  "Ngày giao dịch cuối cùng"),
]


# ── HELPER: GTGD (Giá trị giao dịch tỷ VNĐ) ────────────────

def calc_gtgd(gia_hien_tai, klgd) -> float | None:
    """
    GTGD (tỷ VNĐ) = Giá đóng cửa (đồng) x KLGD (hợp đồng) / 1,000,000,000
    Giả định 1 hợp đồng CW = 1 chứng quyền (chuẩn HOSE).
    """
    if gia_hien_tai is None or klgd is None:
        return None
    try:
        return round(float(gia_hien_tai) * float(klgd) / 1_000_000_000, 6)
    except (ValueError, TypeError):
        return None


# ── XÂY DỰNG DATAFRAME THEO ĐÚNG CỘT ────────────────────────

def build_table1_df(records: list[dict]) -> pd.DataFrame:
    """Xây bảng 'Kết quả giao dịch' từ list record đã enrich (có premium, don_bay...)."""
    rows = []
    for r in records:
        gia = r.get("gia_hien_tai")
        klgd = r.get("klgd")

        # FIX: dùng "is not None" thay vì "or" để tránh lỗi falsy-zero
        # (0 phiên còn lại là giá trị hợp lệ và QUAN TRỌNG - nghĩa là hôm nay
        # là ngày giao dịch cuối cùng - không được để "or" nuốt mất thành None)
        so_phien = r.get("so_phien_con_lai")
        if so_phien is None:
            so_phien = r.get("so_ngay_den_han")

        row = {
            "ma_cw":               r.get("ma_cw"),
            "pct_thay_doi":        r.get("pct_thay_doi"),
            "gia_hien_tai":        gia,
            "klgd":                klgd,
            "gtgd_ty_vnd":         calc_gtgd(gia, klgd),
            "trang_thai_tien":     r.get("trang_thai_tien") or r.get("trang_thai_cw"),
            "premium":             r.get("premium"),
            "don_bay":             r.get("don_bay"),
            "so_phien_con_lai":    so_phien,
            "do_bien_dong_lich_su":r.get("do_bien_dong_lich_su"),
        }
        rows.append(row)

    df = pd.DataFrame(rows)
    col_order = [c[0] for c in TABLE1_COLUMNS]
    df = df[[c for c in col_order if c in df.columns]]
    df = df.rename(columns=dict(TABLE1_COLUMNS))
    return df


def build_table2_df(records: list[dict]) -> pd.DataFrame:
    """Xây bảng 'Thông tin chứng quyền' từ list record gốc từ scraper."""
    rows = []
    for r in records:
        row = {
            "ma_cw":             r.get("ma_cw"),
            "to_chuc_ph_cw":     r.get("to_chuc_ph_cw"),
            "thoi_han":          r.get("thoi_han"),
            "ty_le_chuyen_doi":  r.get("ty_le_chuyen_doi"),
            "gia_phat_hanh":     r.get("gia_phat_hanh"),
            "kl_niem_yet":       r.get("kl_niem_yet"),
            "gia_thuc_hien":     r.get("gia_thuc_hien"),
            "ngay_gd_cuoi_cung": r.get("ngay_gd_cuoi_cung"),
        }
        rows.append(row)

    df = pd.DataFrame(rows)
    col_order = [c[0] for c in TABLE2_COLUMNS]
    df = df[[c for c in col_order if c in df.columns]]
    df = df.rename(columns=dict(TABLE2_COLUMNS))
    return df


# ── ĐỊNH DẠNG EXCEL (giống style PDF: header xanh, viền, căn giữa) ──

HEADER_FILL  = PatternFill(start_color="1F6FB2", end_color="1F6FB2", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
CELL_FONT    = Font(size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

ITM_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # xanh nhạt
OTM_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # đỏ nhạt


def _style_worksheet(ws, df: pd.DataFrame, moneyness_col: str | None = None):
    """Áp style cho 1 worksheet: header, border, auto-width, highlight ITM/OTM."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    moneyness_idx = None
    if moneyness_col and moneyness_col in df.columns:
        moneyness_idx = list(df.columns).index(moneyness_col) + 1

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        if moneyness_idx:
            status_val = ws.cell(row=row_idx, column=moneyness_idx).value
            if status_val == "ITM":
                ws.cell(row=row_idx, column=moneyness_idx).fill = ITM_FILL
            elif status_val == "OTM":
                ws.cell(row=row_idx, column=moneyness_idx).fill = OTM_FILL

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))
        for val in df.iloc[:, col_idx - 1]:
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 35)

    ws.freeze_panes = "A2"


# ── HÀM CHÍNH: XUẤT 2 FILE EXCEL ────────────────────────────

def export_two_excel_files(records_active: list[dict],
                            output_dir: str = ".",
                            report_date: date | None = None) -> tuple[str, str]:
    """
    Xuất 2 file Excel từ list record CW đã enrich (có premium, don_bay...).
    Trả về (đường_dẫn_file1, đường_dẫn_file2).
    """
    if report_date is None:
        report_date = date.today()

    date_str = report_date.strftime("%Y%m%d")

    df1 = build_table1_df(records_active)
    df2 = build_table2_df(records_active)

    fname1 = f"{output_dir}/KetQuaGiaoDich_{date_str}.xlsx"
    fname2 = f"{output_dir}/ThongTinChungQuyen_{date_str}.xlsx"

    with pd.ExcelWriter(fname1, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="KetQuaGiaoDich", index=False, startrow=0)
        ws = writer.sheets["KetQuaGiaoDich"]
        _style_worksheet(ws, df1, moneyness_col="Trạng thái tiền")

    with pd.ExcelWriter(fname2, engine="openpyxl") as writer:
        df2.to_excel(writer, sheet_name="ThongTinChungQuyen", index=False, startrow=0)
        ws = writer.sheets["ThongTinChungQuyen"]
        _style_worksheet(ws, df2)

    return fname1, fname2


if __name__ == "__main__":
    sample_records = [
        {
            "ma_cw": "CACB2510", "pct_thay_doi": "+11.11%", "gia_hien_tai": 500,
            "klgd": 195200, "trang_thai_cw": "ITM", "premium": 5.2, "don_bay": 6.63,
            "so_phien_con_lai": 5, "do_bien_dong_lich_su": 31.5,
            "to_chuc_ph_cw": "SSI", "thoi_han": "12 tháng", "ty_le_chuyen_doi": "2 : 1",
            "gia_phat_hanh": "1,800", "kl_niem_yet": "1,500,000",
            "gia_thuc_hien": "22,500", "ngay_gd_cuoi_cung": "19/06/2026",
        },
        {
            "ma_cw": "CFPT2518", "pct_thay_doi": "+7.1%", "gia_hien_tai": 150,
            "klgd": 162400, "trang_thai_cw": "OTM", "premium": 49.3, "don_bay": 4.54,
            "so_phien_con_lai": 0, "do_bien_dong_lich_su": 45.0,
            "to_chuc_ph_cw": "SSI", "thoi_han": "15 tháng", "ty_le_chuyen_doi": "8.61 : 1",
            "gia_phat_hanh": "2,600", "kl_niem_yet": "1,500,000",
            "gia_thuc_hien": "105,888", "ngay_gd_cuoi_cung": "21/09/2026",
        },
    ]

    f1, f2 = export_two_excel_files(sample_records, output_dir="/tmp")
    print(f"✅ Đã tạo: {f1}")
    print(f"✅ Đã tạo: {f2}")

    df_check = pd.read_excel(f1)
    print(df_check[["Mã", "Số phiên còn lại"]].to_string(index=False))
